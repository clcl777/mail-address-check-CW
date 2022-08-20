import logging
import time
import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
import random
file = 'test.xlsx'
wb1 = openpyxl.load_workbook(file)
ws1 = wb1.worksheets[0]
values = []
for cell in ws1['B']:
    values.append(cell.value)
#始まりの列番号

i = 3
values = values[i-1:]
options = Options()
options.add_argument('--no-sandbox')
#options.add_argument('--headless')
options.add_argument('--disable-gpu')
options.add_argument('--lang=ja-JP')
options.add_argument(
    f'user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/79.0.3945.79 Safari/537.36')
options.add_argument('--user-data-dir=C:\\Users\\tisk0\\AppData\\Local\\Google\\Chrome\\User Data')
options.add_argument('--profile-directory=Profile 4')
driver = webdriver.Chrome(ChromeDriverManager().install(),options=options)
driver.maximize_window()
driver.implicitly_wait(5)
q1 = 'https://www.google.com/search?q=%22'
q2 = '%22'

url = 'info@fujita-jibika.jp'
print(url)
driver.get(q1 + url + q2)
time.sleep(random.random()*2)
current_url = driver.current_url
if 'sorry' in current_url:
    #recapcha必要
    time.sleep(60)
    print('recapchaあり')
    driver.close()
    driver.quit()
    driver = webdriver.Chrome(ChromeDriverManager().install())
    driver.maximize_window()
    driver.implicitly_wait(5)
    driver.get(q1 + url + q2)
    time.sleep(1)
html_txt = driver.page_source
if 'との一致はありません。' in html_txt:
    #存在しない
    print('存在しない')
    ws1.cell(i, 3).value = '×'
else:
    # 存在するかも
    try:
        print('e2')
        url_found_element = driver.find_element_by_css_selector(
            '#rso > div:nth-child(1) > div > div > div.yuRUbf > a')
        url_found = url_found_element.get_attribute("href")
        driver.get(url_found)
        html_found = driver.page_source
        if url in html_found:
            print(url_found)
            ws1.cell(i, 3).value = url_found
            print('存在する')
        else:

            print('存在しない')
            ws1.cell(i, 3).value = '×'
    except:
        print('e3')
        #logging.exception("What is doing when exception happens.")
        #0件表示
        zero_result = driver.find_elements_by_css_selector('#result-stats')
        print('e5')
        if len(zero_result) > 0:
            print(zero_result[0].get_attribute('textContent'))
            marker_kensu = zero_result[0].get_attribute('textContent')
            if '約 0 件' in zero_result[0].get_attribute('textContent'):
                print('存在しない')
                ws1.cell(i, 3).value = '×'
            else:
                print('e6')
                url_found_element = driver.find_element_by_css_selector('#rso > div > div > div > div > div.yuRUbf > a')
                url_found = url_found_element.get_attribute("href")
                print(url_found)
        elif '検索条件と十分に一致する結果が見つかりません。' in html_txt:
            print('ee')
            ws1.cell(i, 3).value = '×'
        else:
            #一致なし
            #ws1.cell(i, 3).value = 'エラー発生'
            print('エラー発生')
            print('e4')


