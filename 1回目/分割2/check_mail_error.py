import logging
import re

import openpyxl
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager

file = '211105　info検証中.xlsx'
print(file)
wb1 = openpyxl.load_workbook(file)
ws1 = wb1.worksheets[1]
i = 5
values_mail = []
values_url = []
values_result = []
for cell in ws1['B']:
    values_mail.append(cell.value)
for cell2 in ws1['C']:
    values_url.append(cell2.value)
for cell3 in ws1['D']:
    values_result.append(cell3.value)
values_mail = values_mail[i:]
values_url = values_url[i:]
values_result = values_result[i:]

#options = Options()
#options.add_argument('--no-sandbox')
#options.add_argument('--headless')
#options.add_argument('--disable-gpu')
#options.add_argument('--lang=ja-JP')
#options.add_argument(
    #f'user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/79.0.3945.79 Safari/537.36')
#options.add_argument('--user-data-dir=C:\\Users\\tisk0\\AppData\\Local\\Google\\Chrome\\User Data')
#options.add_argument('--profile-directory=Profile 4')
#driver = webdriver.Chrome(ChromeDriverManager().install(),options=options)
#driver.maximize_window()
#driver.implicitly_wait(5)

for url, mail, result in zip(values_url, values_mail, values_result):
    print(i)
    if not result == 'エラー':
        i = i + 1
        continue
    '''
    try:
        print(mail)
        driver.get(url)
        html = driver.page_source.encode('utf-8')
        soup = BeautifulSoup(html, "html.parser")
        text_soup = str(soup)
        new_mail = re.findall('[a-zA-Z0-9._-]*' + mail, text_soup)[0]
        print(new_mail)
        ws1.cell(i + 1, 4).value = new_mail
    except:
        print('エラー発生')
        logging.exception("What is doing when exception happens.")
        ws1.cell(i+1, 4).value = 'エラー'
    '''
    ws1.cell(i + 1, 4).value = mail
    #wb1.save(file)

    i = i + 1
wb1.save(file)