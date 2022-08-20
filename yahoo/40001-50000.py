import logging
import time
import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
import random
file = '40001-50000.xlsx'
print(file)
wb1 = openpyxl.load_workbook(file)
ws1 = wb1.worksheets[0]
values = []
for cell in ws1['B']:
    values.append(cell.value)
#始まりの列番号

i = 3
values = values[i-1:]
options = Options()
options.add_argument('--no-sandbox')
#options.add_argument('--headless')
options.add_argument('--disable-gpu')
options.add_argument('--lang=ja-JP')
options.add_argument(
    f'user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/79.0.3945.79 Safari/537.36')
#options.add_argument('--user-data-dir=C:\\Users\\tisk0\\AppData\\Local\\Google\\Chrome\\User Data')
#options.add_argument('--profile-directory=Profile 4')
driver = webdriver.Chrome(ChromeDriverManager().install(),options=options)
driver.maximize_window()
driver.implicitly_wait(5)
q1 = 'https://search.yahoo.co.jp/search?p=%22'
q2 = '%22'
for url in values:
    try:
        print(i)
        print(url)
        driver.get(q1 + url + q2)
        time.sleep(random.random()*2 + 1)
        html_txt = driver.page_source
        if 'ご覧になろうとしているページは現在表示できません。' in html_txt:
            #エラー
            time.sleep(random.random()*2 + 1)
            print('ページエラー')
            driver.close()
            driver.quit()
            driver = webdriver.Chrome(ChromeDriverManager().install())
            driver.maximize_window()
            driver.implicitly_wait(5)
            driver.get(q1 + url + q2)
            time.sleep(random.random()*2 + 1)
        html_txt = driver.page_source
        if '順番も含め完全に一致するウェブページは見つかりませんでした。' in html_txt or 'に一致する情報は見つかりませんでした。' in html_txt:
            # 存在しない
            print('存在しない')
            ws1.cell(i, 3).value = '×'
        else:
            #存在あり
            url_found_elements = driver.find_elements_by_class_name('sw-Card__title')
            nashi = True
            for url_found_element in url_found_elements:
                url_found_element = url_found_element.find_element_by_tag_name('a')
                url_found = url_found_element.get_attribute("href")
                driver.get(url_found)
                html_found = driver.page_source
                if url in html_found:
                    print(url_found)
                    ws1.cell(i, 3).value = url_found
                    print('存在する')
                    nashi = False
                    break
                break
            if nashi:
                print('存在しない')
                ws1.cell(i, 3).value = '×'
    except:
        logging.exception("What is doing when exception happens.")
        print('エラー')
        ws1.cell(i, 3).value = 'エラー'

    i = i + 1
    time.sleep(3*random.random() + 1)
    wb1.save(file)
